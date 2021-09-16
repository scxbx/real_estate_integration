import datetime
import json
import os
from tkinter import filedialog, ttk
from tkinter import messagebox
import openpyxl
import tkinter as tk
import re

# class parcel:
#     a_shi_bie_ma = ''
#     b_yao_su_dai_ma = ''
#     c_xian_qu_dai_ma = ''
#     d_di_ji_qu = ''
#     e_di_ji_zi_qu = ''
#     f_zong_di_shun_xu_hao = ''
#     g_tu_di_suo_you_quan_lei_xing = ''
#     h_zong_di_te_zheng_ma = ''
#     i_zong_di_dai_ma = ''
#     j_zhen = ''
#     k_cun = ''
#     l_zu_ming = ''
#     m_zong_di_ming_cheng = ''
#     n_bu_dong_chan_bian_hao = ''
#     o_zong_di_zuo_luo = ''
#     p_suo_you_quan_quan_li_ren = ''
#     q_quan_li_ren_ming_cheng = ''
#     r_quan_li_ren_lei_xing = ''
#     s_zheng_jian_zhong_lei = ''
#     t_zheng_jian_bian_hao = ''
#     u_dian_hua_hao_ma = ''
#     v_tong_xun_di_zhi = ''
#     w_suo_you_quan_zong_di_dai_ma = ''
import find_missing_images
import ocr_id

complete_time_dict = {}


def sum_dict(a, b):
    """
        相同的键，值相加
    """
    temp = dict()
    # dict_keys类似set； | 并集
    for key in a.keys() | b.keys():
        temp[key] = sum([d.get(key, 0) for d in (a, b)])
    return temp


def get_max_row(sheet):
    """
        获取工作表的最大行数
    """
    max_row = len([row for row in sheet if not all([cell.value is None for cell in row])])
    return max_row


def get_from_info(aip):
    """
        根据【地籍区】代码填写【镇】
        根据【地籍区+地籍子区】代码填写【村】

        :param aip: 配号申请信息表
        :return 字典（地籍子区后六位 -> [村委会行政区编码, 村委会行政区名称, 地级子区名称]
    """
    aib = openpyxl.load_workbook(aip, data_only=True)
    ais = aib.active
    max_row = get_max_row(ais)
    # print('max row', max_row)

    ai_dict = {}
    for cur_row in range(2, max_row + 1):
        # print(cur_row, ais.cell(cur_row, 4).value)
        # D 乡镇代码
        d = ais.cell(cur_row, 4).value
        # E　乡镇名称
        e = ais.cell(cur_row, 5).value
        # F 村委会行政区编码
        f = ais.cell(cur_row, 6).value
        # G 村委会行政区名称
        g = ais.cell(cur_row, 7).value
        # H 地籍子区代码
        h = ais.cell(cur_row, 8).value
        # I  地级子区名称
        i = ais.cell(cur_row, 9).value

        # print(d, e, h, i)

        if d is not None and e is not None:
            ai_dict[d] = e
            if h is not None and i is not None and f is not None and g is not None:
                ai_dict[h[-6:]] = [f, g, i]
    # print(ai_dict)
    return ai_dict


def fill_parcel_table(ppp, ai_dict, para_list, json_path, images_path, use_ocr=False, ):
    """
    填单个宗地属性表

    :param ppp: parcel_property_path宗地属性路径的简写
    :param ai_dict: 从“配号申请信息表”中获取的数据的字典
    :param para_list: 从gui输入的数据
    :param json_path: 处理后的json文件夹路径
    :param images_path: 02电子档案
    :param use_ocr: 是否使用ocr 默认为Fasle
    :return:
    """
    # D 地籍区 like 200

    d = ''
    # E 地籍子区 like 002
    e = ''
    # J 镇
    j = ''
    # K 村 对应配号申请信息表中的”地籍子区名称“ [f,g,i]中的i
    k = ''
    # X 所属行政村代码
    x = ''

    # 权利人名称相同
    list_identical_name = []
    msg_identical_name = '权利人姓名相同'
    # 无对应身份证号
    list_no_id = []
    msg_no_id = '无对应身份证号'
    # 房屋超出175平方米
    list_house_larger = []
    msg_house_larger = '房屋超出175平方米'
    # 宗地面积超出限制
    list_parcel_larger = []
    msg_parcel_larger = '宗地面积超出限制'
    # 占地面积大于宗地面积，请核实
    list_cover_larger_parcel = []
    msg_cover_larger_parcel = '占地面积大于宗地面积，请核实'
    # 身份证错误
    list_id_error = []
    msg_id_error = '身份证错误，请核实'

    # parcel property book
    ppb = openpyxl.load_workbook(ppp, data_only=True)
    pps = ppb['宗地属性表']
    max_row = get_max_row(pps)
    p_area_dict = {}  # 宗地面积的字典
    c_area_dict = {}  # 建筑占地面积的字典
    f_dict = {}  # 宗地顺序号
    field_l = ''
    name_id_dict = {}
    image_name_dict = get_jpg_filenames(images_path)

    def get_cell(col):
        return pps.cell(cur_row, col).value

    def get_cell_letter(col):
        col_num = convert_to_number(col)
        return get_cell(col_num)

    def update_cell(col, value):
        pps.cell(cur_row, col).value = value

    def update_cell_letter(col, value):
        col_num = convert_to_number(col)
        return update_cell(col_num, value)

    def log_with_row(msg):
        print('行数：{}\t{}'.format(cur_row, msg))

    if use_ocr:
        name_id_dict = get_parcel_id_dict(json_path)
    # print('name_id_dict, name_id_dict')

    for cur_row in range(2, max_row + 1):
        # print(pps.cell(cur_row, 4).value)
        d = pps.cell(cur_row, 4).value  # 地籍区
        e = pps.cell(cur_row, 5).value  # 地籍子区
        j = ai_dict.get(d)
        d_list = ai_dict.get(d + e)
        if d_list is not None:
            k = d_list[2]

        # print('f', pps.cell(cur_row, 6).value)
        # print('e', e)
        # print('j', j)
        # print('k', k)

        pps.cell(cur_row, 10).value = j
        pps.cell(cur_row, 11).value = k

        # 根据输入的村小组填写【组名】
        field_l = para_list[0]
        pps.cell(cur_row, 12).value = field_l

        # 根据【镇】、【村】、【组名】填写【宗地坐落】 O
        o = '陵水黎族自治县' + j + k + field_l
        pps.cell(cur_row, 15).value = o

        # 根据【镇】、【村】、【组名】填写【所有权权利人】 【组名】+"经济合作社农民集体"
        p = field_l + '经济合作社农民集体'
        pps.cell(cur_row, 16).value = p

        # 根据照片/填写【证件编号】 注意：无论权利人数量多少，都只填入1-1的身份证号，需要人工调整
        if use_ocr:
            i = pps.cell(cur_row, 9).value
            t = name_id_dict.get(i) if name_id_dict.get(i) is not None else ''
            # print('name_id_dict.get(i)', name_id_dict.get(i))
            # print('t', t)
            pps.cell(cur_row, 20).value = t
            if not is_valid_idcard(t):
                q = pps.cell(cur_row, 17).value  # Q: 权利人姓名
                f = pps.cell(cur_row, 6).value  # F: 宗地顺序号
                list_id_error.append("{} {} {}".format(q, f, t))
            # print("id error {} {} {}".format(q, f, t))
            # print(list_id_error)
        # 根据【村】填写【所属行政村代码】
        if d_list is not None:
            x = d_list[0]
            pps.cell(cur_row, 24).value = x

            # 根据【村】填写【所属行政村名称】
            y = d_list[1]
            pps.cell(cur_row, 25).value = y

        # 填写【权利类型】 填写 6
        z = 6
        pps.cell(cur_row, 26).value = z

        # 填写【权利性质】    填写 203
        aa = 203
        pps.cell(cur_row, 27).value = aa

        # 填写【权力设定方式】	填写 2
        ab = 2
        pps.cell(cur_row, 28).value = ab

        # 填写【批准用途】	填写 0702 注意不能填写成702
        ac = '0702'
        pps.cell(cur_row, 29).value = ac

        # 根据【宗地面积】AJ 填写【批准面积】AI	填写【宗地面积】，保留两位小数
        aj = pps.cell(cur_row, 36).value
        ai = "%.2f" % float(aj)
        pps.cell(cur_row, 35).value = ai

        # 检查【宗地面积】
        f = pps.cell(cur_row, 6).value  # F: 宗地顺序号
        q = pps.cell(cur_row, 17).value  # Q: 权利人姓名
        an = pps.cell(cur_row, 40).value  # AN: 建筑占地面积
        p_area_dict = sum_dict(p_area_dict, {q: float(aj)})
        c_area_dict = sum_dict(c_area_dict, {q: float(an)})

        # 如果 权利人名称 -> 宗地顺序号 这个字典中已存在和本次循环相同的权力人名称，将宗地顺序号合并后赋给value
        if f_dict.get(q) is None:
            f_dict[q] = f
        else:
            f_dict[q] = f + ' ' + f_dict[q]

        # 无对应身份证号（T）
        t = pps.cell(cur_row, 20).value
        if t is None:
            list_no_id.append(q + ' ' + f)

        # 填写宗地面积单位 设为1
        al = 1
        pps.cell(cur_row, 38).value = al

        # 根据【权利人姓名 Q】填写【共有权利人情况AZ 52】
        # 一个人时不填写，2个人时填写【权利人姓名】，【权利人类型 1】，【证件种类 1】，【证件编号 T】；
        # 若人数过多（>=3），则填写第一个【权利人姓名】+【等X人】。
        comma = '，'
        if comma in q:
            q_list = q.split(comma)
            q_size = len(q_list)

            t_list = []
            t_size = 0

            if t is not None:
                if comma in t:
                    t_list = t.split(comma)
                    t_size = len(t_list)

            if q_size >= 3:
                az = '{}等{}人'.format(q_list[0], q_size)
            elif q_size == 2 and t_size == 2:
                az = '{}，1，1，{}，{}，1，1，{}'.format(q_list[0], t_list[0], q_list[1], t_list[1])
            else:
                az = ''
                print('填写【共有权利人情况】出现异常！ 行数：', cur_row)

            pps.cell(cur_row, 52).value = az

        # 填写宅基地来源   填写1
        bm = 1
        pps.cell(cur_row, 65).value = bm

        # 填写【宅基地主要占地类型】	填写 3
        bn = 3
        pps.cell(cur_row, 66).value = bn

        # 填写【比例尺】	填写 1：2000
        bp = '1:2000'
        pps.cell(cur_row, 68).value = bp

        # 填写【调查审核人】	填写 方荣
        cd = '方荣'
        pps.cell(cur_row, 82).value = cd

        # 填写【权籍调查单位】	填写 江西省煤田地质局测绘大队
        cf = '江西省煤田地质局测绘大队'
        pps.cell(cur_row, 84).value = cf

        # 根据输入/输入表格填写【调查员BY 77】、【调查日期BZ 78】、【测量人CB 80】、【测量日期CC 81】、【调查审核日期CE 83】
        # 根据输入或提前输入的表格填写以上字段值
        # 日期格式为8位数字，不用符号
        by = para_list[1]
        bz = para_list[2]
        cb = para_list[3]
        cc = para_list[4]
        ce = para_list[5]

        pps.cell(cur_row, 77).value = by
        pps.cell(cur_row, 78).value = bz
        pps.cell(cur_row, 80).value = cb
        pps.cell(cur_row, 81).value = cc
        pps.cell(cur_row, 83).value = ce

        # 填写状态 BV 74    填入固定值1
        pps.cell(cur_row, 74).value = '1'

        # 填写实际用途 AD 30  填入固定值0702
        pps.cell(cur_row, 30).value = '0702'

        # 根据【权利人姓名Q17】填写【项目名称CN92】	填写【权利人姓名】+“农房”
        q = get_cell(17)
        cn = q if q is not None else ''
        cn = cn + '农房'
        update_cell(92, cn)

        # 填写【项目性质CO93】	填写 农民自有房屋建设
        update_cell(93, '农民自有房屋建设')

        # 根据【镇J10】填写【权利人邮编CP94】	"提蒙/提蒙乡-填写572435 本号/本号镇-填写572434"
        j = get_cell(10)
        if '提蒙' in j:
            tmp = 572435
        elif '本号' in j:
            tmp = 572434
        else:
            tmp = ''
            log_with_row('镇既不是本号也不是提蒙，填入空值')
        update_cell(94, tmp)

        # 填写【国家CQ95】	填写 142
        update_cell(95, 142)

        # 填写【户籍所在省市CR96】	填写 460000
        update_cell(96, 460000)

        # 填写【是否核查CV100】	填写 1
        update_cell(100, 1)

        # 填写【申请人类型CX102】	填写 1
        update_cell(102, 1)

        # 填写【土地来源CY103】	填写 1
        update_cell(103, 1)

        # 根据/02电子档案内的照片填写【宗地归档CU99】
        # 填写文件夹内所有照片的文件名（含后缀名），按顺序（1-1、2-X、4-X）,每个名字用英文分号 ; 隔开
        i = get_cell(9)  # 宗地代码I9
        if image_name_dict is None:
            in_image_file_dict = None
        else:
            in_image_file_dict = image_name_dict.get(i)
        image_str = ''
        if in_image_file_dict is not None:
            l1 = in_image_file_dict.get('l1')
            if l1 is not None:
                for item in l1:
                    image_str += item + ';'
            l2 = in_image_file_dict.get('l2')
            if l2 is not None:
                for item in l2:
                    image_str += item + ';'
            l4 = in_image_file_dict.get('l4')
            if l4 is not None:
                sort_img_name(l4)
                for item in l4:
                    image_str += item + ';'
        update_cell(99, image_str)

        # 根据/02电子档案内的4-X照片填写【图片信息 CW101】
        # 注意：当对应单元格已有数据时，不改变原有数据
        # 填写文件夹内4-X照片的文件名（含后缀名），按顺序（4-1、4-2、...）,每个名字用英文分号 ; 隔开。
        # 若文件夹内不存在4-X照片，填写4-1.jpg,0,0,0,0。
        cw = get_cell(101)
        if cw is None:
            img_info = ''
            if in_image_file_dict is not None and in_image_file_dict.get('l4') is not None:
                l4 = in_image_file_dict.get('l4')
                if len(l4) == 0:
                    img_info = '4-1.jpg,0,0,0,0'
                else:
                    sort_img_name(l4)
                    for item in l4:
                        img_info += item + ',0,0,0,0;'
            update_cell(101, img_info)

        # 填写【土地权属来源证明材料】(BW列)	填写 见权属来源证明等附件
        update_cell_letter('BW', '见权属来源证明等附件')

        # 填写【是否存在争议】（CZ列）	填写 0
        update_cell_letter('cz', 0)

        # 填写【审核意见】（DA列）	填写 经核实该宗地权属调查、地籍测量成果合格。
        update_cell_letter('da', '经核实该宗地权属调查、地籍测量成果合格')

        # 填写【权属调查记事】（BX列）	填写 该宗地经实地调查测量核实，权属界线清楚，符合规划，权属来源材料合法，四邻确认无异议。
        update_cell_letter('bx', '该宗地经实地调查测量核实，权属界线清楚，符合规划，权属来源材料合法，四邻确认无异议')

        # 填写【状态】（BV列）	填写 1
        update_cell_letter('bv', 1)

        # 根据配号信息表/【所属行政村代码】、【宗地代码】填写【地籍号】(BJ列)
        # "填写【所属行政村代码 d_list[0]】+（【宗地代码 i】最后四位-{输入地籍号差值 para_list[6]}）+000
        if d_list is not None:
            i = get_cell_letter('i')
            if para_list[6] != '':
                bj = d_list[0] + str(int(i[-4:]) - int(para_list[6])).zfill(4) + '000'
                update_cell_letter('bj', bj)

        # 填写【不动产测量记事】（CA列）	填写 该宗地经现场核实，界址点位置、四至无争议，界址点采用解析法（或其他方法）测量，
        # 宗地面积按解析法计算，宗地图按地籍规范测绘。
        update_cell_letter('ca', '该宗地经现场核实，界址点位置、四至无争议，界址点采用解析法（或其他方法）测量，宗地面积按解析法计算，宗地图按地籍规范测绘。')

    # print(p_area_dict)
    # print(c_area_dict)

    for key in p_area_dict:
        p_area = p_area_dict.get(key, 0)
        c_area = c_area_dict.get(key, 0)
        if p_area >= 175.01:
            if abs(p_area - c_area) < 0.01:
                print(key, '房屋超出175平方米', p_area, c_area)
                list_house_larger.append(key + ' ' + f_dict[key])
            else:
                print(key, '宗地面积超出限制', p_area, c_area)
                list_parcel_larger.append(key + ' ' + f_dict[key])
        if c_area > p_area:
            print(key, '占地面积大于宗地面积，请核实', p_area, c_area)
            list_cover_larger_parcel.append(key + ' ' + f_dict[key])
        value_f_dict = f_dict[key]
        if ' ' in value_f_dict:
            list_identical_name.append(key + ' ' + value_f_dict)

    ppb.save(ppp)
    ppb.close()

    # 将检查结果写入“宗地属性检查.txt”
    exam_images_file_name = r"结果\宗地属性检查.txt"
    if os.path.exists(exam_images_file_name):
        f_exam = open(exam_images_file_name, 'w')
    else:
        f_exam = open(exam_images_file_name, "x")
    write_list(f_exam, list_identical_name, msg_identical_name)
    write_list(f_exam, list_no_id, msg_no_id)
    write_list(f_exam, list_house_larger, msg_house_larger)
    write_list(f_exam, list_parcel_larger, msg_parcel_larger)
    write_list(f_exam, list_cover_larger_parcel, msg_cover_larger_parcel)
    write_list(f_exam, list_id_error, msg_id_error)
    f_exam.close()
    return


"""
    填房屋属性表
    hpb: house property path 填房屋属性表路径
"""


def fill_house_table(hpp, para_list):
    """
    填房屋属性表

    :param hpp: house property path 填房屋属性表路径
    :param para_list: 从gui输入的数据数组
    """
    # house property book
    hpb = openpyxl.load_workbook(hpp, data_only=True)
    hps = hpb['房屋属性表']
    max_row = get_max_row(hps)
    comma = '，'
    for c_row in range(2, max_row + 1):
        # 宗地代码为None的一行不需要做任何处理
        if hps.cell(c_row, 2).value is None or hps.cell(c_row, 2).value == '':
            # print('宗地代码为None。row:', c_row)
            continue

        def update_cell(col, value):
            """
            更新房屋属性表中单元格信息

            :param col: 要更改的单元格的列
            :param value: 新的值
            """
            hps.cell(c_row, col).value = value

        def get_cell(col):
            """
            获取房屋属性表单元格的值

            :param col: 要获取的单元格的列
            :return: 单元格的值
            """
            return hps.cell(c_row, col).value

        # 填写【定着物代码D4】	填写 F
        d = 'F'
        update_cell(4, d)

        # 填写【权利人类型J10】	填写 1
        j = 1
        update_cell(10, j)

        # 填写【证件种类K11】	填写 1
        k = 1
        update_cell(11, k)

        # 根据【房屋坐落G7】填写【住址N14】	填写【房屋坐落】
        g = get_cell(7)
        update_cell(14, g)

        # 根据【房屋所有权人I9】填写【共有情况O15】	一个人时填写 单独所有 ，多个人时填写 共同共有
        i = get_cell(9)
        if i is not None:
            if len(i.split(comma)) > 1:
                o = '共同所有'
            else:
                o = '单独所有'
            update_cell(15, o)

            # 根据【共有情况O15】填写【共有方式P16】	单独所有 填写 0 ；共同共有 填写 1
            if o == '共同所有':
                p = 1
            elif o == '单独所有':
                p = 0
            else:
                p = 0
                print('共有情况既不为共同所有，也不为单独所有，共有方式填入默认值0。row:', c_row)
            update_cell(16, p)

            # 根据【房屋所有权人I9】填写【项目名称Q17】	【房屋所有权人】+农房

            q = i + '农房'
            update_cell(17, q)

        # 填写【产别S19】	填写 30
        s = 30
        update_cell(19, s)

        # 填写【用途T20】	填写 10
        t = 10
        update_cell(20, t)

        # 填写【用途名称U21】	填写 住宅
        u = '住宅'
        update_cell(21, u)

        # 填写【规划用途V22】	填写 10
        v = 10
        update_cell(22, 10)

        # 填写【产权来源AL38】	填写 100
        al = 100
        update_cell(38, al)

        # 填写【墙体归属AM39 AN40 AO41 AP42】	东南西北 原为批量填写10，现改为空值才填写"10"
        am, an, ao, ap = 10, 10, 10, 10
        wall_list = [am, an, ao, ap]
        for iw in range(0, 4):
            if get_cell(iw + 39) is None:
                update_cell(iw + 39, wall_list[iw])

        # 填写【是否核查AR44】	填写 1
        ar = 1
        update_cell(44, ar)

        # 填写【调查意见AT46】	填写 房籍调查结果合格。
        at = '房籍调查结果合格。'
        update_cell(46, at)

        # 根据输入/输入表格填写【调查员AU47 para_list[1]】、【调查日期AV48 para_list[2]】	填写与宗地表相同的【调查员】、【调查日期】
        au = para_list[1]
        av = para_list[2]
        update_cell(47, au)
        update_cell(48, av)

        # 将竣工时间填入complete_time_dict 宗地代码 B2 -> 竣工时间 AG33
        if get_cell(2) not in complete_time_dict:
            complete_time_dict[get_cell(2)] = get_cell(33)
        elif complete_time_dict[get_cell(2)] is not None:
            # print('get_cell(33)', get_cell(33))
            # print('get_cell(2)', get_cell(2))
            # print('complete_time_dict[get_cell(2)]', complete_time_dict[get_cell(2)])
            later_date = get_later_date(get_cell(33), complete_time_dict[get_cell(2)])
            complete_time_dict[get_cell(2)] = later_date

    hpb.save(hpp)
    hpb.close()
    return


def get_later_date(date1, date2):
    """
    从两个日期字符串中获取较新的那个字符串

    :param date1: 日期字符串1
    :param date2: 日期字符串2
    :return: 较新的那个字符串
    """
    if date1 is None or date2 is None:
        return ''
    d1_l = date1.split('-')
    d2_l = date2.split('-')
    # print(d1_l)
    # print(d2_l)
    d1 = datetime.datetime(int(d1_l[0]), int(d1_l[1]), int(d1_l[2]))
    d2 = datetime.datetime(int(d2_l[0]), int(d2_l[1]), int(d2_l[2]))
    if d1 > d2:
        return date1
    else:
        return date2


def fill_announce_table(ppp, sample_announce_path, out_announce_path):
    """
    填写公示表

    :param ppp: 房屋属性表
    :param sample_announce_path: 公示表模板
    :param out_announce_path: 填写完的公示表的保存路径
    """
    ppb = openpyxl.load_workbook(ppp, data_only=True)
    pps = ppb['宗地属性表']
    pp_max_row = get_max_row(pps)

    announce_book = openpyxl.load_workbook(sample_announce_path, data_only=True)
    a_sheet = announce_book.active
    announce_max_row = get_max_row(a_sheet)
    a_sheet['A1'] = '陵水黎族自治县{}{}{}不动产权籍调查结果公示表'.format(pps['J2'].value, pps['K2'].value, pps['L2'].value)

    # print('complete_time_dict', complete_time_dict)

    for a_row in range(4, announce_max_row + 1):
        p_row = a_row - 2

        def update_cell(col, value):
            a_sheet.cell(a_row, col).value = value

        def update_a_with_p(a_col, p_col):
            a_sheet.cell(a_row, a_col).value = pps.cell(p_row, p_col).value

        # 填写【占地批准用途 Ｅ５】	填写 农村宅基地
        update_cell(5, '农村宅基地')

        # 预编宗地代码        B2
        update_a_with_p(2, 9)
        # 权利人名称          C3     Q17
        update_a_with_p(3, 17)
        # 权利人身份证件编号   D4      T20
        update_a_with_p(4, 20)
        # 占地批准用途        E5      AC29
        update_a_with_p(5, 29)
        # 宗地面积	        F6      AJ36
        area = pps.cell(p_row, 36).value
        if area is not None:
            update_cell(6, "%.2f" % float(area))

        # 房屋建筑面积	    G7      AM39
        update_a_with_p(7, 39)
        # 房屋建筑占地面积	    H8      AN40
        update_a_with_p(8, 40)
        # 房屋建筑时间	    I9      房屋属性表 AG33
        code = a_sheet.cell(a_row, 2).value
        # print("code", code)
        # print("complete_time_dict.get(code)", complete_time_dict.get(code))
        if complete_time_dict.get(code) is not None:
            update_cell(9, complete_time_dict.get(code))
        # 备注               J10     不填
    announce_book.save(out_announce_path)
    announce_book.close()
    ppb.close()


def write_list(my_file, my_list, my_str):
    """
    将错误信息写入文本文档

    :param my_file: 保存的文档的路径
    :param my_list: 要写入的数组
    :param my_str: 要写入的标题
    :return:
    """
    if len(my_list) != 0:
        my_file.write('【' + my_str + '】\n')
        my_file.write('\n'.join(my_list) + '\n\n')


def simple_entry(master, label_text):
    frame = tk.Frame(master)
    label = tk.Label(frame, text=label_text)
    label.pack(side=tk.LEFT)
    entry = tk.Entry(frame)
    entry.pack(side=tk.RIGHT)
    frame.pack()
    return entry


def simple_combobox(master, label_text, value_list):
    frame = tk.Frame(master)
    label = tk.Label(frame, text=label_text)
    label.pack(side=tk.LEFT)
    com = ttk.Combobox(frame, state="readonly", values=value_list)  # #创建下拉菜单
    com.pack(side=tk.RIGHT)  # #将下拉菜单绑定到窗体
    com.current(0)  # #设定下拉菜单的默认值为第0个

    def xFunc(event):
        print(com.get())  # #获取选中的值方法1

    com.bind("<<ComboboxSelected>>", xFunc)  # #给下拉菜单绑定事件
    frame.pack()
    return com


IDCARD_REGEX = '[1-9][0-9]{14}([0-9]{2}[0-9X])?'


def is_valid_idcard(idcard):
    if isinstance(idcard, int):
        idcard = str(idcard)

    if not re.match(IDCARD_REGEX, idcard):
        return False

    items = [int(item) for item in idcard[:-1]]

    # 加权因子表
    factors = (7, 9, 10, 5, 8, 4, 2, 1, 6, 3, 7, 9, 10, 5, 8, 4, 2)

    # 计算17位数字各位数字与对应的加权因子的乘积
    copulas = sum([a * b for a, b in zip(factors, items)])

    # 校验码表
    ckcodes = ('1', '0', 'X', '9', '8', '7', '6', '5', '4', '3', '2')

    return ckcodes[copulas % 11].upper() == idcard[-1].upper()


def get_jpg_filenames(path):
    if path == '':
        return None
    out_dict = {}
    dir_list = os.listdir(path)
    for my_dir in dir_list:
        f_path = os.path.join(path, my_dir)
        if os.path.isdir(f_path):
            list1 = []
            list2 = []
            list4 = []
            file_list = os.listdir(f_path)
            for file in file_list:
                f_file = os.path.join(f_path, file)
                if os.path.isfile(f_file):
                    if file.endswith('.jpg'):
                        type_num = find_missing_images.check_image_type(file)
                        if type_num == '1':
                            list1.append(file)
                        elif type_num == '2':
                            list2.append(file)
                        elif type_num == '4':
                            list4.append(file)
            in_dict = {'l1': list1, 'l2': list2, 'l4': list4}
            out_dict[my_dir[:19]] = in_dict
    return out_dict


def sort_img_name(a_list):
    """
    将元素均为‘4-x房屋照片.jpg的数组按照数字x排序

    :param a_list: 要排序的数组
    """

    def takSecond(item):
        string = re.match('(.*)-(.*)房屋照片.jpg', item).group(2)
        return int(string)

    a_list.sort(key=takSecond)


def convert_to_number(letter, column_a=1):
    """
    字母列号转数字
    columnA: 你希望A列是第几列(0 or 1)? 默认1
    return: int
    """
    ab = '_ABCDEFGHIJKLMNOPQRSTUVWXYZ'
    letter0 = letter.upper()
    w = 0
    for _ in letter0:
        w *= 26
        w += ab.find(_)
    return w - 1 + column_a


def next_day(date_string):
    from datetime import datetime
    from datetime import timedelta
    datetime_obj = datetime.strptime(date_string, '%Y-%m-%d').date()
    tomorrow = datetime_obj + timedelta(days=1)
    return str(tomorrow)


def json_to_dict(path):
    with open(r'默认表格\config.txt', encoding='utf-8') as f:
        content = json.loads(f.read())
        return content


class App(tk.Tk):
    """
    使用tk的简单gui
    """
    # path_input_parcel = r'C:\Users\sc\Documents\飞行者科技\房地一体\测试数据\成果2-填写了身份证及其他信息的宗地属性表.xlsx'
    # path_input_house = r'C:\Users\sc\Documents\飞行者科技\房地一体\测试数据\46坡田仔\08过程数据\房屋属性表.xlsx'
    path_input_info = r'默认表格\陵水县配号申请信息表.xlsx'
    path_input_announce = r'默认表格\成果模板-公示表.xlsx'
    path_input_json = ''
    path_input_image = ''
    path_input_house = ''
    path_input_parcel = ''

    # village_group = '老宗园下'
    # investigator = '黄开博'
    # invest_date = '2021-03-13'
    # surveyor = '曾广深'
    # survey_date = '2021-03-14'
    # review_date = '2021-03-15'

    def __init__(self):
        super().__init__()

        # 宗地属性表路径
        parcel_path = tk.StringVar()
        tk.Label(self, text="宗地属性表路径（必填）:").pack()
        tk.Entry(self, textvariable=parcel_path, width=30).pack()

        def selectPath_parcel():
            self.path_input_parcel = filedialog.askopenfilename(title='请选择宗地属性表', filetypes=[('Excel', '.xlsx')])
            parcel_path.set(self.path_input_parcel)
            # print(path_input_parcel)

        tk.Button(self, text="路径选择", command=selectPath_parcel).pack()

        # 房屋属性表
        house_path = tk.StringVar()
        tk.Label(self, text="房屋属性表路径（必填）:").pack()
        tk.Entry(self, textvariable=house_path, width=30).pack()

        def selectPath_house():
            self.path_input_house = filedialog.askopenfilename(title='房屋属性表', filetypes=[('Excel', '.xlsx')])
            house_path.set(self.path_input_house)

        tk.Button(self, text="路径选择", command=selectPath_house).pack()

        # image文件夹
        image_path = tk.StringVar()
        tk.Label(self, text="“02电子档案”文件夹路径（选填）:").pack()
        tk.Entry(self, textvariable=image_path, width=30).pack()

        def selectPath_image():
            self.path_input_image = filedialog.askdirectory()
            image_path.set(self.path_input_image)

        tk.Button(self, text="路径选择", command=selectPath_image).pack()

        # 配号申请信息表
        info_path = tk.StringVar()
        tk.Label(self, text="配号申请信息表路径:").pack()
        tk.Entry(self, textvariable=info_path, width=30).pack()

        def selectPath_info():
            self.path_input_info = filedialog.askopenfilename(title='配号申请信息表', filetypes=[('Excel', '.xlsx')])
            info_path.set(self.path_input_info)

        tk.Button(self, text="路径选择", command=selectPath_info).pack()

        # 公示表
        announce_path = tk.StringVar()
        tk.Label(self, text="公示表路径:").pack()
        tk.Entry(self, textvariable=announce_path, width=30).pack()

        def selectPath_announce():
            self.path_input_announce = filedialog.askopenfilename(title='公示表', filetypes=[('Excel', '.xlsx')])
            announce_path.set(self.path_input_announce)

        tk.Button(self, text="路径选择", command=selectPath_announce).pack()

        # json文件夹
        json_path = tk.StringVar()
        tk.Label(self, text="json文件夹路径（暂时不填）:").pack()
        tk.Entry(self, textvariable=json_path, width=30).pack()

        def selectPath_json():
            self.path_input_json = filedialog.askdirectory()
            json_path.set(self.path_input_json)

        tk.Button(self, text="路径选择", command=selectPath_json).pack()

        # 组名    para_list[0]
        entry_village_group = simple_entry(self, '组名')
        # 调查员   para_list[1]
        # entry_investigator = simple_entry(self, '调查员')
        config_dict = json_to_dict(r'默认表格\config.txt')
        com_investigator = simple_combobox(self, '调查员', config_dict.get('调查员'))
        # 调查日期  para_list[2]
        entry_invest_date = simple_entry(self, '调查日期(yyyy-mm-dd)')
        # 测量人   para_list[3]
        # entry_surveyor = simple_entry(self, '测量人')
        com_surveyor = simple_combobox(self, '测量人', config_dict.get('测量人'))
        # 测量日期  para_list[4]
        # entry_survey_date = simple_entry(self, '测量日期(yyyy-mm-dd)')
        # 调查审核日期    para_list[5]
        # entry_review_date = simple_entry(self, '调查审核日期(yyyy-mm-dd)')
        # 地籍号差值     para_list[6]
        entry_land_num_dif = simple_entry(self, '地籍号差值')

        def btn_generator_CallBack():
            village_group = entry_village_group.get()
            investigator = com_investigator.get()
            invest_date = entry_invest_date.get()
            surveyor = com_surveyor.get()
            survey_date = invest_date

            try:
                review_date = next_day(invest_date)
            except ValueError as e:
                tk.messagebox.showerror(title="错误", message="日期格式错误")
                print(e.args)
                return
            land_num_dif = entry_land_num_dif.get()

            para_list = [village_group,
                         investigator,
                         invest_date,
                         surveyor,
                         survey_date,
                         review_date,
                         land_num_dif]

            # messagebox.showinfo("Hello Python", para_list)

            # print(os.getcwd())

            if self.path_input_parcel != '' and self.path_input_house != '':
                out_path = r'结果\公示表.xlsx'
                fill_parcel_table(self.path_input_parcel,
                                  get_from_info(self.path_input_info),
                                  para_list, self.path_input_json,
                                  self.path_input_image)
                print('宗地属性表填写完毕')
                fill_house_table(self.path_input_house, para_list)
                print('房屋属性表填写完毕')
                fill_announce_table(self.path_input_parcel, self.path_input_announce, out_path)
                print('公示表填写完毕')
                print('三个表同时填写完毕')
            else:
                messagebox.showerror('错误', '请选择宗地属性表和房屋属性表')

        B = tk.Button(self, text="写入", command=btn_generator_CallBack)
        B.pack()


def get_parcel_id_dict(path):
    ocr_dict = ocr_id.read_jsons(path)
    name_id_dict = {}
    print('ocr_dict', ocr_dict)
    for item in ocr_dict:
        name = item.get('宗地号')
        id_num = item.get('权利人证件编号')
        name_id_dict[name] = id_num
    return name_id_dict


if __name__ == '__main__':
    # print('Sad Tuesday')
    # in_path = tk.filedialog.askopenfilename()
    # print(in_path)
    # fill_single(r'C:/Users/sc/Documents/飞行者科技/房地一体/测试数据/成果2-填写了身份证及其他信息的宗地属性表.xlsx', get_from_info(r'C:/Users/sc/Documents/飞行者科技/房地一体/测试数据/陵水县配号申请信息表.xlsx'))
    # get_from_info(r'C:/Users/sc/Documents/飞行者科技/房地一体/测试数据/陵水县配号申请信息表.xlsx')

    app = App()
    app.mainloop()
