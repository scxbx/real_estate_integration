import os

import openpyxl
import pandas as pd
from tkinter import filedialog

from parcel_property_sheet import get_max_row


def derive_relation(relation_dict, leader_relation, member_relation, is_male, is_elder):
    # 保存关系的字典（权利人在户口本上的关系, 成员在户口本上的关系) -> 成员相对权利人上的关系
    # 需要判断年龄和性别时，value为list [大男, 小男, 大女, 小女]
    # my_dict = {('长子', '户主'): ['父亲', '父亲', '母亲', '母亲'],
    #            ('子', '妻'): '母亲',
    #            ('长子', '妻'): '母亲',
    #            ('长子', '长子'): '兄弟',
    #            ('长子', '长女'): '长子, 长女',
    #            ('长子', '父亲'): '祖父',
    #            ('长子', '母亲'): '祖母',
    #            ('长子', '妹妹'): '长子, 妹妹'}

    new_relation = relation_dict.get((leader_relation, member_relation))
    if new_relation is not None:
        if type(new_relation) is str:
            return new_relation
        else:
            if is_male:
                if is_elder:
                    return new_relation[0]
                else:
                    return new_relation[1]
            else:
                if is_elder:
                    return new_relation[2]
                else:
                    return new_relation[3]


class Transformer:
    error_list = []
    relation_dict_path = ''
    relation_list_path = ''
    error_path = ''
    form_path = ''
    out_path = ''

    def __init__(self, relation_dict_path, relation_list_path, error_path, form_path, out_path):
        self.error_path = error_path
        self.relation_dict_path = relation_dict_path
        self.relation_list_path = relation_list_path
        self.form_path = form_path
        self.out_path = out_path

    def transform(self):
        my_dict = self.get_dict_from_csv()
        wb = openpyxl.load_workbook(self.form_path)
        ws = wb.active
        max_row = get_max_row(ws)

        start_row = 3
        for row_c in range(start_row, max_row + 1):
            zong_code = ws.cell(row_c, 1).value
            leader = ws.cell(row_c, 2).value
            family_count = ws.cell(row_c, 3).value
            leader_relation = ''
            leader_index = 0

            def log(msg, row):
                complete_msg = "{}: {} {} {}".format(row, zong_code[-3:], leader, msg)
                self.error_list.append(complete_msg)

            if family_count is not None:
                for i in range(family_count):
                    if leader == ws.cell(row_c + i, 4).value:
                        leader_relation = ws.cell(row_c + i, 5).value
                        leader_index = i

                if leader_relation is None:
                    print('row {}: leader_relation is None'.format(row_c))
                    log('权利人关系为空', row_c)

                elif leader_relation == '户主':
                    # print('row {}: need no transformation'.format(row_c))
                    pass
                else:
                    print('row {}: leader_relation: {}'.format(row_c, leader_relation))
                    for i in range(family_count):
                        member = ws.cell(row_c + i, 4).value
                        member_relation = ws.cell(row_c + i, 5).value

                        def log(msg, row):
                            complete_msg = "{}: {} {} {} {}".format(row, zong_code[-3:], leader, member, msg)
                            self.error_list.append(complete_msg)

                        # print('do some relational transformation here')

                        if not self.pass_pre_check(member_relation):
                            print('row {}: 关系 {} 不在已有列表中'.format(row_c + i, member_relation))
                            log("关系 {} 不在已有列表中".format(member_relation), row_c + i)
                        if i == leader_index:
                            ws.cell(row_c + i, 5).value = '户主'
                            print('row {}: special change ({}, {}) to 户主'.format(row_c + i,
                                                                                 leader_relation,
                                                                                 member_relation))
                        else:
                            member_age = ws.cell(row_c + i, 14).value
                            leader_age = ws.cell(row_c + leader_index, 14).value
                            member_gender = ws.cell(row_c + i, 8).value
                            if member_age is not None and leader_age is not None:
                                is_elder = member_age > leader_age
                            else:
                                is_elder = True
                                print('row {}: 错误年龄 {} {} 当作大'.format(row_c + i, leader_age, member_age))
                                log('错误年龄 {} {} 当作成员比权利人大'.format(leader_age, member_age), row_c + i)
                            if member_gender == '男':
                                is_male = True
                            elif member_gender == '女':
                                is_male = False
                            else:
                                is_male = True
                                print('row {}: 错误性别 {} 当作男'.format(row_c + i, member_gender))
                                log('错误性别 {} 成员当作男'.format(member_gender), row_c + i)
                            new = derive_relation(my_dict, leader_relation, member_relation, is_male, is_elder)
                            if new is not None:
                                ws.cell(row_c + i, 5).value = new
                                print(
                                    'row {}: change ({}, {}) to {}'.format(row_c + i, leader_relation,
                                                                           member_relation,
                                                                           new))
                            else:
                                print(
                                    'row {}: relation ({}, {}) not in dict'.format(row_c + i, leader_relation,
                                                                                   member_relation))
                                log('关系 ({}, {}) 不在关系字典中'.format(leader_relation, member_relation), row_c + i)
        # new = os.path.splitext(self.form_path)[0] + 'new.xlsx'
        wb.save(self.out_path)
        print('已生成新挂接表：{}'.format(self.out_path))
        self.write_error_txt()
        print('已生成关系转化反馈：{}'.format(self.error_path))

    def write_error_txt(self):
        with open(self.error_path, encoding='utf-8', mode='w+') as f:
            f.write('行数 宗地号 权利人 成员 错误信息\n')
            for item in self.error_list:
                f.write(item + '\n')

    def pass_pre_check(self, member_relation):
        # relation_list = ["户主", "父亲", "母亲",
        #                  "子", "儿子", "长子", "次子",
        #                  "二子", "三子", "四子", "五子",
        #                  "女", "女儿", "长女", "次女", "二女",
        #                  "三女", "四女", "五女", "兄", "弟", "兄弟",
        #                  "哥哥", "弟弟", "姐", "妹", "姐妹", "姐姐",
        #                  "妹妹", "儿媳", "妻", "妻子", "弟媳", "配偶",
        #                  "嫂子", "孙", "孙女", "孙子", "外孙子", "外孙",
        #                  "外孙女", "侄子", "侄女", "外甥", "外甥女"]
        relation_list = self.get_list_from_txt()
        if member_relation in relation_list:
            return True
        else:
            return False

    def get_list_from_txt(self):
        """
        从txt中获取关系列表。txt文件中，关系用逗号隔开
        :return: 关系列表
        """
        relation_list = []
        with open(self.relation_list_path, encoding='utf-8', mode='r') as f:
            lines = f.readlines()
            for line in lines:
                for item in line.split(','):
                    if item != '\n':
                        relation_list.append(item.strip())
        return relation_list

    def get_dict_from_csv(self):
        """
        将csv文件的数据转为dict

        :param self: self
        :return: 字典 (权利人在户口本上的关系, 成员在户口本上的关系) -> 成员相对权利人上的关系
        """

        df = pd.read_csv(self.relation_dict_path)
        relation_dict = {}
        for i in range(len(df)):
            # print(df.iloc[i][0], df.iloc[i][1], df.iloc[i][2])
            if ',' in df.iloc[i][2]:
                # print('', ' in df.iloc[i][2]', df.iloc[i][2])
                new_list = df.iloc[i][2].split(',')
                relation_dict[(df.iloc[i][0], df.iloc[i][1])] = new_list
            else:
                relation_dict[(df.iloc[i][0], df.iloc[i][1])] = df.iloc[i][2]

        return relation_dict


if __name__ == '__main__':
    print("Hello Monday")
    f_path = filedialog.askopenfilename()

    tran = Transformer(relation_dict_path=r'关系转换/relation.csv',
                       relation_list_path=r'关系转换/relation_list.txt',
                       error_path=r'结果/关系转化反馈.txt',
                       form_path=f_path,
                       out_path=r'结果/新挂接表.xlsx')
    tran.transform()

    input('按任意键结束')
