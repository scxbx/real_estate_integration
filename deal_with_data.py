# coding=utf-8

# coding=gbk
import os
import threading

import check_id
import json
import openpyxl
import jsonToExcel
import tkinter as tk
from tkinter import filedialog, messagebox

# df = pd.read_json('1.json')

# print(df)

import ocr_id
from parcel_property_sheet import simple_entry

msg_error = []


def write_into_excel(in_path, out_path):
    wb = openpyxl.load_workbook(in_path, data_only=True)
    ws = wb['Sheet1']

    row = 3
    for i in range(len(jsonToExcel.member_name)):
        members_num = len(jsonToExcel.member_name[i])
        for j in range(members_num):
            def fill_and_merge(my_col, my_list):
                if j == 0:
                    # print('row', row)
                    ws.cell(row, my_col).value = my_list[i][j]
                    # print('row2', row)
                    ws.merge_cells(start_row=row, start_column=my_col, end_row=row + members_num - 1, end_column=my_col)

            id_number = jsonToExcel.idcard[i][j]

            check_id.check_id(id_number, row, jsonToExcel.zong_num[i][j][-3:], jsonToExcel.leader_name[i][j])

            # 宗地号 A 1
            fill_and_merge(1, jsonToExcel.zong_num)

            # 户主 B2
            fill_and_merge(2, jsonToExcel.leader_name)
            # 家庭总人数 C3
            fill_and_merge(3, jsonToExcel.fam)
            # 户内成员姓名    D4
            ws.cell(row, 4).value = jsonToExcel.member_name[i][j]
            # 与户主关系     E5
            ws.cell(row, 5).value = jsonToExcel.relation[i][j]
            # 性别    H 8
            ws.cell(row, 8).value = get_gender_from_id(id_number, row)
            # 出生日期  J 10
            ws.cell(row, 10).value = get_birthday_from_id(id_number)
            # 身份证号  M 13
            ws.cell(row, 13).value = id_number
            # 年龄    N 14
            ws.cell(row, 14).value = get_age_from_id(id_number)
            # 备注    R

            row += 1

    wb.save(out_path)
    wb.close()

    clear_data()
    print('写入{}完成'.format(out_path))


def clear_data():
    jsonToExcel.member_name = []
    jsonToExcel.relation = []
    jsonToExcel.leader_name = []
    jsonToExcel.idcard = []
    jsonToExcel.zong_num = []
    jsonToExcel.leader_id = []
    jsonToExcel.family_num = 0
    jsonToExcel.fam = []


def get_age_from_id(id_number):
    if id_number is None:
        return ''
    if len(id_number) != 18:
        return ''

    year = id_number[6:10]
    if year.isdigit():
        year = int(year)
        age = 2021 - year
        return age
    else:
        return ''


def get_birthday_from_id(id_number):
    if id_number is None:
        return ''
    if len(id_number) != 18:
        return ''
    year = id_number[6:10]
    month = id_number[10:12]
    day = id_number[12:14]
    return '{}年{}月{}日'.format(year, month, day)


def get_gender_from_id(id_number, row):
    if id_number is None:
        return ''
    if len(id_number) != 18:
        return ''
    gender_num = id_number[-2]
    if gender_num.isdigit():
        gender_num = int(gender_num)
        if gender_num % 2 == 1:
            return '男'
        else:
            return '女'
    else:
        msg_error.append('第{}行身份证倒数第二位不为数字，无法识别男女'.format(row))
        return ''


def read_jsons_to_list(path):
    files = os.listdir()
    my_list = []
    for file in files:
        if '.json' in file:
            full_path = os.path.join(path, file)
            with open(full_path, 'r') as f:
                temp = json.loads(f.read())
                my_list.append(temp)
                # print(temp)

    return my_list


def write_error_msg(alist, apath):
    with open(apath, 'w+') as f:
        for i in alist:
            f.write(i + '\n')


def switch_off(btn_list):
    for btn in btn_list:
        if btn['state'] == 'normal':
            btn['state'] = 'disabled'


def switch_on(btn_list):
    for btn in btn_list:
        if btn['state'] == 'disabled':
            btn['state'] = 'normal'


class App(tk.Tk):
    """
        使用tk的简单gui。用于：
        1. 通过ocr获取由每个家庭的字典组成的数组，并保存过程中的json文件
        2. 通过保存的json文件，直接生成由每个家庭的字典组成的数组
        3. 通过上述数组，生成挂接表
    """

    list_families = []
    sample_path = ''
    out_path = ''
    json_dir = ''
    village_name = ''
    btn_ocr = None
    btn_local = None
    btn_local_json = None
    btn_write = None
    btn_write_choose = None

    def __init__(self, sample_path, out_path):
        super().__init__()
        self.title('飞行者OCR识别填表工具')
        self.geometry('360x{}'.format(self.winfo_reqheight()))
        self.sample_path = sample_path
        self.out_path = out_path

        self.entry_village_group = simple_entry(self, '村小组名称')

        self.btn_ocr = tk.Button(self, width="20", text="OCR识别", pady=4,
                                 command=lambda: self.thread_it(self.ocr_to_list))
        self.btn_local = tk.Button(self, width="20", text="本地数据", pady=4, command=self.json_to_list)
        self.btn_local_json = tk.Button(self, width="20", text="本地数据json", pady=4, command=self.direct_json_to_list)
        self.btn_write = tk.Button(self, width="20", text="写入excel", pady=4,
                                   command=self.list_to_excel, state='disabled')
        self.btn_write_choose = tk.Button(self, width="20", pady=4, text="选取文件夹并写入excel",
                                          command=self.list_to_excel_choose)

        self.btn_ocr.pack()
        self.btn_local.pack()
        self.btn_local_json.pack()
        self.btn_write.pack()
        self.btn_write_choose.pack()

        # WM_DELETE_WINDOW 不能改变，这是捕获命令
        self.protocol('WM_DELETE_WINDOW', self.on_closing)

        # # menu
        # menuBar = tk.Menu(self)
        # helpMenu = tk.Menu(menuBar, tearoff=0)
        # helpMenu.add_command(label="关于")
        # helpMenu.add_command(label="帮助")
        # menuBar.add_cascade(label='Help', menu=helpMenu)
        # self.config(menu=menuBar)
        # text1 = tk.Text(self, width=65, height=30)
        #
        # scroll = tk.Scrollbar()
        # # 将滚动条填充
        # scroll.pack(side=tk.RIGHT, fill=tk.Y)  # side是滚动条放置的位置，上下左右。fill是将滚动条沿着y轴填充
        #
        # # 将滚动条与文本框关联
        # scroll.config(command=text1.yview)  # 将文本框关联到滚动条上，滚动条滑动，文本框跟随滑动
        # text1.config(yscrollcommand=scroll.set)  # 将滚动条关联到文本框
        #
        # text1.pack()

    # def new_line(self):
    #     tk.Label(self, text='\n').pack()

    def on_closing(self):
        if messagebox.askokcancel("退出", "是否确认退出程序?"):
            self.destroy()

    def ocr_to_list(self):
        # before ocr, disable the ocr button
        btn_to_switch_off = [self.btn_ocr, self.btn_write_choose, self.btn_local, self.btn_local_json, self.btn_write]
        btn_to_switch_on = [self.btn_ocr, self.btn_write_choose, self.btn_local, self.btn_local_json]
        switch_off(btn_to_switch_off)
        self.village_name = self.entry_village_group.get()
        self.list_families = ocr_id.get_family(self.village_name, True)
        self.json_dir = self.list_families[1] if self.list_families is not None else ''
        self.btn_write['state'] = 'normal' if self.list_families is not None else 'disabled'
        switch_on(btn_to_switch_on)

    def json_to_list(self):
        self.village_name = self.entry_village_group.get()
        self.list_families = ocr_id.get_family(self.village_name, False)
        self.json_dir = self.list_families[1] if self.list_families is not None else ''
        self.btn_write['state'] = 'normal' if self.list_families is not None else 'disabled'

    def direct_json_to_list(self):
        self.village_name = self.entry_village_group.get()
        self.list_families = ocr_id.get_family_from_json_folder(self.village_name)
        self.json_dir = self.list_families[1] if self.list_families is not None else ''
        self.btn_write['state'] = 'normal' if self.list_families is not None else 'disabled'

    def list_to_excel(self, is_choose=False):
        print(self.json_dir)
        if self.json_dir == '' and not is_choose:
            print('使用本功能（写入excel）前，请先使用 ocr识别、本地数据或本地数据json')
            return
        path_time = os.path.join('结果', self.json_dir)
        if not os.path.exists(path_time):
            os.makedirs(path_time)
        # my_json_path = r'C:\Users\sc\PycharmProjects\real_estate_integration\2021-08-19 09-31-12'
        if is_choose:
            to_print = '选择文件夹并写入excel: 请选择本软件[json]文件夹下对应的子文件夹'
            my_json_path = filedialog.askdirectory(title=to_print, initialdir='json')
            if my_json_path == '':
                print(to_print)
                return
        else:
            my_json_path = os.path.join('json', self.json_dir)
        print(my_json_path)
        jsonToExcel.info_basic_out(my_json_path)
        if os.path.exists(r'refsrc.xlsx'):
            jsonToExcel.match_ref(r'refsrc.xlsx', os.path.join(path_time, '家庭成员核查反馈.txt'))

        # jsonToExcel.relation_solu()

        write_into_excel(self.sample_path, os.path.join(path_time, '挂接表.xlsx'))
        check_id.write_id_error_file(os.path.join(path_time, '家庭成员核查反馈.txt'))

    def list_to_excel_choose(self):
        self.list_to_excel(True)

    @staticmethod
    def thread_it(func, *args):
        t = threading.Thread(target=func, args=args)
        t.setDaemon(True)  # 守护--就算主界面关闭，线程也会留守后台运行（不对!）
        t.start()  # 启动
        # t.join()          # 阻塞--会卡死界面！


if __name__ == '__main__':
    # a_dict = get_col_data()

    # list_m = read_jsons_to_list(r'C:\Users\sc\PycharmProjects\real_estate_integration')
    # print(list_m)
    # path = r'C:\Users\sc\PycharmProjects\real_estate_integration\json'
    # jsonToExcel.info_basic_out(path)
    # print(jsonToExcel.idcard)
    #
    # write_into_excel(r'C:\Users\sc\Documents\飞行者科技\房地一体\sample.xlsx',
    #                  r'C:\Users\sc\Documents\飞行者科技\房地一体\挂接表out.xlsx')

    in_path = r'默认表格\挂接表模板.xlsx'
    out_path = r'结果\挂接表'
    app = App(in_path, out_path)
    app.mainloop()

    # path = r'C:\Users\sc\PycharmProjects\real_estate_integration\2021-08-18 17-52-08'
    # jsonToExcel.info_basic_out(path)
    # write_into_excel(in_path, out_path)
