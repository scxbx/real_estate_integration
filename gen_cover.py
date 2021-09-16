import openpyxl

import os


def get_max_row(sheet):
    """
        获取工作表的最大行数
    """
    max_row = len([row for row in sheet if not all([cell.value is None for cell in row])])
    return max_row


def multi_cover(path, info, new_path):
    wb = openpyxl.load_workbook(path, data_only=True)
    # ws = wb.active
    # max_row = get_max_row(ws)

    source = wb.active
    for i in range(len(info)):
        # ws = wb.create_sheet(str(i + 1))
        single_cover(wb, info[i], str(i + 1))

    wb.remove(source)
    wb.save(new_path)
    wb.close()
    return


def single_cover(wb, one_cover_list, sheet_name):
    # wb = openpyxl.load_workbook(path, data_only=True)
    # ws = wb.active
    # max_row = get_max_row(ws)
    source = wb.active
    ws = wb.copy_worksheet(source)

    ws['B5'].value = one_cover_list[0]
    ws['E6'].value = one_cover_list[1]
    ws['B7'].value = one_cover_list[2]
    ws['E7'].value = one_cover_list[3]
    ws['B8'].value = one_cover_list[4]

    ws.title = sheet_name
    return


def get_info(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb['Sheet1']

    max_row = get_max_row(ws)

    info_list = []

    for row in range(2, max_row + 1):
        one_row_list = []
        for col in range(1, 6):
            one_row_list.append(ws.cell(row, col).value)
        info_list.append(one_row_list)

    wb.close()
    return info_list


if __name__ == '__main__':
    info_path = os.path.join(os.getcwd(), 'info.xlsx')
    full_info = get_info(info_path)

    # single_cover(r'C:\Users\sc\Documents\飞行者科技\房地一体\曾广深\cover.xlsx', full_info[0], 'test.xlsx')
    cover_path = os.path.join(os.getcwd(), 'cover.xlsx')
    results_path = os.path.join(os.getcwd(), 'results')
    # for i in range(len(full_info)):
    #     single_cover(cover_path, full_info[i], os.path.join(results_path, str(i + 1) + '.xlsx'))
    new_file_path = os.path.join(results_path, 'result.xlsx')
    multi_cover(cover_path, full_info, new_file_path)
