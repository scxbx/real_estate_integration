import os
from tkinter import filedialog
from openpyxl import load_workbook
from parcel_property_sheet import get_max_row


def update(archives_path, xlsx_path):
    dict_name_to_zong = {}
    pl = os.listdir(archives_path)
    for p in pl:
        zong = p[:19]
        name = p[19:]
        dict_name_to_zong[name] = zong

    # print('dict', dict_name_to_zong)

    wb = load_workbook(xlsx_path)
    ws = wb.active
    max_row = get_max_row(ws)

    for row in range(3, max_row + 1):
        name = ws.cell(row, 2).value
        if name is not None:
            # print(name, 'is not None')
            zong = dict_name_to_zong.get(name)
            if zong is not None:
                ws.cell(row, 1).value = zong
                # print('name', name, 'zong', zong)
            else:
                print('找不到权利人 {} 对应的宗地号。可能是权利人名称错误'.format(name))

    wb.save(xlsx_path)
    wb.close()
    print('替换完成')


if __name__ == '__main__':
    a_p = filedialog.askdirectory(title='选择文件夹 02电子档案')
    x_p = filedialog.askopenfilename(title='选择挂接表', filetypes=[('Excel', '*.xlsx')])
    update(a_p, x_p)
    # print('469028200204JC00002卓亚文'[:19], '469028200204JC00002卓亚文'[19:])
