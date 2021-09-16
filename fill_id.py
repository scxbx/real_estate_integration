import openpyxl
import pandas as pd
from tkinter import filedialog

from parcel_property_sheet import get_max_row


def get_zong_id_dict(_path_in):
    df = pd.read_excel(_path_in, header=1)
    df = df.fillna(method="ffill")
    # print(df.iloc[0:20, 0:3])

    zong_id_dict = {}
    for index, row in df.iterrows():
        # print(index, type(row), row['宗地号'], row['户主'], row['户内成员姓名'], row['身份证号'])
        if row['户主'] == row['户内成员姓名']:
            zong_id_dict[row['宗地号']] = row['身份证号']

    return zong_id_dict


def fill(path_, dict_):
    # df = pd.read_excel(path_, header=2)
    # for index, row in df.iterrows():
    #     if dict_.get(row['宗地代码']) is not None:
    #         row['证件编号'] = dict_.get(row['宗地代码'])
    wb = openpyxl.load_workbook(path_)
    ws = wb.active
    max_row = get_max_row(ws)
    for i in range(1, max_row+1):
        print('ws.cell(i, 9).value', ws.cell(i, 9).value)
        print('dict_.get(ws.cell(i, 9).value)', dict_.get(ws.cell(i, 9).value))
        if dict_.get(ws.cell(i, 9).value) is not None:
            ws.cell(i, 20).value = ws.cell(i, 9).value

    wb.save(path_)


if __name__ == '__main__':
    print('Hello Friday')
    # path_in = r'C:\Users\sc\PycharmProjects\real_estate_integration\结果\坡田仔\挂接表.xlsx'
    # path_out = r'C:\Users\sc\Documents\飞行者科技\房地一体\测试数据\backup\成果2-填写了身份证及其他信息的宗地属性表 - 副本.xlsx'
    path_in = filedialog.askopenfilename(title='选择挂接表', filetypes=[('Excel', '*.xlsx')])
    path_out = filedialog.askopenfilename(title='选择宗地属性表', filetypes=[('Excel', '*.xlsx')])

    my_dict = get_zong_id_dict(path_in)
    print(my_dict)
    fill(path_out, my_dict)

